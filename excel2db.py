import os
import argparse
import psycopg2
from openpyxl import load_workbook
from psycopg2 import sql

class Excel2Posrgresql:

    def __init__(self, fileName: str, host: str, database: str, table: str, user: str, password:str, sheetname: str = None, requiredfield: str = None):
        filename = os.path.abspath(args.fileName)
        self.filename = filename
        self.host = host
        self.database = database
        self.table = table
        self.user = user
        self.password = password
        self.sheetname = sheetname
        self.requiredfield = requiredfield

    def ReadFromExcelToDatabase(self):
        work_book = load_workbook(self.filename)
        if self.sheetname == None:
            self.sheet = work_book.active
        else:
            try:
                self.sheet = work_book[self.sheetname]
            except Exception as exception:
                raise exception

        header_cells = self.GetHeader()
        columns = self.CheckDatabase()
        check = self.CheckField(header_cells, columns)
        if check:
            self.InsertToTable(header_cells)
    
    def CheckField(self, header_cells, columns):
        if(self.requiredfield != None):
            for col in self.sheet.iter_cols(min_row=header_cells[0].row+1,max_row=self.sheet.max_row):
                i = 0
                for cell in col:
                    if (header_cells[i].value in self.requiredfield.split(",")):
                        if(cell.value == None):
                            raise Exception("Required cell of column cannot be empty")
                i+=1
        if(header_cells):
            for cell in header_cells:
                print(cell.value, end=",")
                if cell.value != None:
                    if(cell.value.lower() in columns):
                        pass
                    else:
                        raise Exception("Columns in Excel not matched in database")
        return True

    def Connect(self):
        # Connect to PostgreSQL
        try:
            connection = psycopg2.connect(
                host=self.host,
                user=self.user,
                database=self.database,
                password=self.password
            )
            connection.autocommit = True
            return connection
        except Exception as exception:
            raise exception

    def CheckDatabase(self):
        connection = self.Connect()
        with connection.cursor() as cursor:
            list_tables = "SELECT table_name FROM information_schema.tables WHERE table_catalog = lower(%s) AND table_name = lower(%s);"
            cursor.execute(list_tables, (self.database,self.table,))
            result = cursor.fetchone()
            if(not result):
                raise Exception("Table Not Found")
            else:
                list_columns = "SELECT column_name FROM information_schema.columns WHERE table_name = %s;"
                cursor.execute(list_columns, (self.table,))
                results = cursor.fetchall()
                columns = []
                for column in results:
                    columns.append(column[0].lower())
                return columns

    def InsertToTable(self, header_cells):
        values = []
        for cell in header_cells:
            values.append(cell.value.lower())
        connection = self.Connect()
        with connection.cursor() as cursor:
            # Skip header self.header_cells[0].row+1 and None value
            for row in self.sheet.iter_rows(min_row=header_cells[0].row+1, min_col=header_cells[0].column, max_col=header_cells[-1].column, values_only=True):
                insert_query = sql.SQL("""INSERT INTO {} ({}) VALUES ({})""").format(
                    sql.Identifier(self.table),
                    sql.SQL(', ').join(map(sql.Identifier, values)),
                    sql.SQL(', ').join(sql.Placeholder() * len(header_cells))
                )
                print(row)
                cursor.execute(insert_query, row)
                        
    # Since headers cannot be None or empty and should have data and must be at first row we can have this below code
    def GetHeader(self):
        rows = tuple(self.sheet.rows)
        max_len_row = 0
        row_header_cells = []
        for row in rows:
            current_row_length = 0
            for cell in row:
                if (cell.value != None):
                    current_row_length += 1
            if max_len_row < current_row_length:
                max_len_row = current_row_length
                row_header_cells.append(row)

        if(len(row_header_cells) != 1):
            raise Exception("Headers cannot be identified")
        
        start_cell_col = None
        start_cell_row = None
        end_cell_row = None
        end_cell_col = None
        length = 0
        none = 0
        i = 0
        for cell in row_header_cells[0]:
            if(cell.value != None):
                length += 1
                if(length == 1):
                    start_cell_col = cell.column
                    start_cell_row = cell.row
            else:
                none += 1
                if(1 < length):
                    i += 1
            if(length == (len(row_header_cells[0]) - none)):
                end_cell_col = cell.column - i 
                end_cell_row = cell.row
            
        row_reference = self.sheet.iter_rows(min_row=start_cell_row, min_col=start_cell_col ,max_row=end_cell_row,max_col=end_cell_col)
        header = next(row_reference, None)
        """
        for e in self.sheet.iter_rows(min_row=start_cell_row, max_row=self.sheet.max_row, min_col=start_cell_col):
            for cell in e:
                print(cell.value, end=",")
            print()
        """
        return header

    def Print(self):
        for row in tuple(self.sheet.rows):
            for cell in row:
                print(cell.value, end=",")
            print()

    @staticmethod
    def ArgumentParser():
        parse = argparse.ArgumentParser(description="Insert Excel Data From Table To Postgresql database.")
        parse.add_argument("-f", "--fileName", required=True, type=str, help="Excel file name")
        parse.add_argument("-s", "--sheetname", required=False, type=str, help="Sheet name(Optional) if not specified active sheet will be selected.")
        parse.add_argument("-H", "--host", required=True, type=str, help="Hostname for PostgreSQL to create connection")
        parse.add_argument("-u", "--username", required=True, type=str, help="Userame for Database")
        parse.add_argument("-p", "--password", required=True, type=str, help="Password for Database")
        parse.add_argument("-d", "--database", required=True, type=str, help="Database")
        parse.add_argument("-t", "--table", required=True, type=str, help="Table")
        parse.add_argument("-m", "--mandatoryfield", required=False, type=str, help="Mandatoryfield are headers that it's data cannot be null(sperated by commas)")
        args=parse.parse_args()
        return args

if __name__ == "__main__":
    args = Excel2Posrgresql.ArgumentParser()
    o = Excel2Posrgresql(args.fileName,args.host,args.database,args.table,args.username,args.password,args.sheetname,args.mandatoryfield)
    o.ReadFromExcelToDatabase()